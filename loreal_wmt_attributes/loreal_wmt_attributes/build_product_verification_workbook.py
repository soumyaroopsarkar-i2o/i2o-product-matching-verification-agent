import csv
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


BASE = Path(__file__).parent
AMAZON_CSV = BASE / "loreal_amazon_match.csv"
WALMART_CSV = BASE / "loreal_walmart_match.csv"
OUT_XLSX = BASE / "lorealpi_product_verification_input.xlsx"
OUT_CSV = BASE / "lorealpi_product_verification_input.csv"

X_SOURCE_END = 15
Y_TARGET_END = 32

SOURCE_COLUMNS = [
    ("Source_ASIN", "asin"),
    ("Source_UPC", "EACH UPC"),
    ("Source_URL", "amazon_link"),
    ("Source_Brand", "amazon_brand"),
    ("Source_Title", "amazon_title"),
    ("Source_Category", "amazon_category_path"),
    ("Source_Rating", "amazon_rating"),
    ("Source_Ratings_Total", "amazon_ratings_total"),
    ("Source_Price", "amazon_price"),
    ("Source_Currency", "amazon_currency"),
    ("Source_Seller", "amazon_seller"),
    ("Source_Search_Alias", "amazon_search_alias"),
    ("Source_Feature_Bullets", "amazon_feature_bullets"),
    ("Source_Description", "amazon_description"),
    ("Source_Image_URL", "amazon_image"),
]

TARGET_COLUMNS = [
    ("Target_Platform", None),
    ("Target_Item_ID", "walmart_item_id"),
    ("Target_Product_ID", "walmart_product_id"),
    ("Target_URL", "walmart_link"),
    ("Target_Brand", "walmart_brand"),
    ("Target_Title", "walmart_title"),
    ("Target_Model", "walmart_model"),
    ("Target_Type", "walmart_type"),
    ("Target_Category", "walmart_category_path"),
    ("Target_Rating", "walmart_rating"),
    ("Target_Ratings_Total", "walmart_ratings_total"),
    ("Target_Price", "walmart_price"),
    ("Target_Currency", "walmart_currency"),
    ("Target_Seller", "walmart_seller"),
    ("Target_Description", "walmart_description"),
    ("Target_Ingredients", "walmart_ingredients"),
    ("Target_Image_URL", "walmart_image"),
]

OUTPUT_COLUMNS = ["Match Status", "Match Justification"]
HEADERS = [c[0] for c in SOURCE_COLUMNS] + [c[0] for c in TARGET_COLUMNS] + OUTPUT_COLUMNS


def clean(value):
    return (value or "").strip()


def truthy(value):
    return clean(value).lower() in {"true", "1", "yes", "y"}


def amazon_url(asin):
    return f"https://www.amazon.com/dp/{asin}" if asin else ""


def load_amazon_by_upc():
    with AMAZON_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return {
            clean(row.get("EACH UPC")): row
            for row in csv.DictReader(f)
            if clean(row.get("EACH UPC"))
        }


def source_value(source, target, field):
    if field == "asin":
        return clean(source.get("asin")) or clean(source.get("ASIN")) or clean(target.get("ASIN"))
    if field == "amazon_link":
        asin = clean(source.get("asin")) or clean(source.get("ASIN")) or clean(target.get("ASIN"))
        return clean(source.get("amazon_link")) or amazon_url(asin)
    return clean(source.get(field))


def target_value(target, field):
    if field is None:
        return "walmart.com"
    return clean(target.get(field))


def build_rows():
    amazon_by_upc = load_amazon_by_upc()
    rows = []
    skipped_no_source_asin = 0
    skipped_no_target_url = 0

    with WALMART_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        for target in csv.DictReader(f):
            if not truthy(target.get("matched")):
                continue

            upc = clean(target.get("EACH UPC"))
            source = amazon_by_upc.get(upc, {})
            source_asin = source_value(source, target, "asin")
            target_url = target_value(target, "walmart_link")
            if not source_asin:
                skipped_no_source_asin += 1
                continue
            if not target_url:
                skipped_no_target_url += 1
                continue

            row = [source_value(source, target, field) for _, field in SOURCE_COLUMNS]
            row.extend(target_value(target, field) for _, field in TARGET_COLUMNS)
            row.extend(["", ""])
            rows.append(row)

    return rows, skipped_no_source_asin, skipped_no_target_url


def write_csv(rows):
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def set_dimensions(ws):
    widths = {
        "A": 16,
        "B": 16,
        "C": 42,
        "D": 18,
        "E": 55,
        "F": 35,
        "G": 12,
        "H": 15,
        "I": 12,
        "J": 12,
        "K": 22,
        "L": 18,
        "M": 55,
        "N": 55,
        "O": 42,
        "P": 16,
        "Q": 16,
        "R": 18,
        "S": 42,
        "T": 18,
        "U": 55,
        "V": 22,
        "W": 18,
        "X": 35,
        "Y": 12,
        "Z": 15,
        "AA": 12,
        "AB": 12,
        "AC": 22,
        "AD": 55,
        "AE": 45,
        "AF": 42,
        "AG": 18,
        "AH": 55,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 36


def style_input_sheet(ws):
    source_fill = PatternFill("solid", fgColor="D9EAF7")
    target_fill = PatternFill("solid", fgColor="E2F0D9")
    output_fill = PatternFill("solid", fgColor="FFF2CC")
    border = Border(bottom=Side(style="thin", color="B7B7B7"))

    for cell in ws[1]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if cell.column <= X_SOURCE_END:
            cell.fill = source_fill
        elif cell.column <= Y_TARGET_END:
            cell.fill = target_fill
        else:
            cell.fill = output_fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_dimensions(ws)

    url_cols = [3, 15, 19, 32]
    for col in url_cols:
        letter = get_column_letter(col)
        for cell in ws[f"{letter}"][1:]:
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    dv = DataValidation(type="list", formula1='"Exact,Equivalent,Not a Match"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"AG2:AG{ws.max_row}")


def add_config_sheet(wb, row_count):
    ws = wb.create_sheet("Skill_Config")
    rows = [
        ("Purpose", "Input sheet for product-verification skill."),
        ("Main sheet", "Verification_Input"),
        ("Original product", "Amazon Source"),
        ("Matched product", "Walmart Target"),
        ("X", X_SOURCE_END),
        ("Y", Y_TARGET_END),
        ("Source range", "Columns 1-15 / A:O"),
        ("Target range", "Columns 16-32 / P:AF"),
        ("Output columns", "Columns 33-34 / AG:AH"),
        ("Rows to verify", row_count),
        ("Suggested prompt", f"verify the product matches in the file {OUT_CSV.name} using the product-verification skill. Use X={X_SOURCE_END} and Y={Y_TARGET_END}. Columns 1-{X_SOURCE_END} are the original Amazon Source product. Columns {X_SOURCE_END + 1}-{Y_TARGET_END} are the matched Walmart Target product."),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_column_map_sheet(wb):
    ws = wb.create_sheet("Column_Map")
    ws.append(["Section", "Column", "Source field"])
    for name, field in SOURCE_COLUMNS:
        ws.append(["Source / Amazon", name, field])
    for name, field in TARGET_COLUMNS:
        ws.append(["Target / Walmart", name, field or "constant: walmart.com"])
    for name in OUTPUT_COLUMNS:
        ws.append(["Verification Output", name, "to be filled by product-verification skill"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EDEDED")


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification_Input"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    style_input_sheet(ws)
    add_config_sheet(wb, len(rows))
    add_column_map_sheet(wb)
    wb.save(OUT_XLSX)


def main():
    rows, skipped_no_source_asin, skipped_no_target_url = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(f"rows={len(rows)}")
    print(f"columns={len(HEADERS)}")
    print(f"X={X_SOURCE_END}")
    print(f"Y={Y_TARGET_END}")
    print(f"skipped_no_source_asin={skipped_no_source_asin}")
    print(f"skipped_no_target_url={skipped_no_target_url}")
    print(f"created_on={datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%d-%m-%Y')}")
    print(f"xlsx={OUT_XLSX}")
    print(f"csv={OUT_CSV}")


if __name__ == "__main__":
    main()
