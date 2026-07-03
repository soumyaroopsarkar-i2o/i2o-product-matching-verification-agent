from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


def compact_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_text_any_encoding(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def safe_text(value: Any, *, limit: int | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if limit is not None and len(text) > limit:
        return text[:limit].rstrip() + "..."
    return text


def normalize_upc(value: Any) -> str:
    text = safe_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return re.sub(r"\D+", "", text)


def parse_price(value: Any) -> float | None:
    text = safe_text(value)
    if not text:
        return None
    lowered = text.lower()
    if any(token in lowered for token in ("free", "n/a", "nan", "none", "null")):
        return None
    if re.search(r"\d\s*[-–]\s*\d", text):
        return None
    cleaned = re.sub(r"[^0-9.,-]", "", text)
    if not cleaned or cleaned in {"-", ".", ","}:
        return None
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        price = float(cleaned)
    except ValueError:
        return None
    if price <= 0:
        return None
    return price


def load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(row_iter)
    except StopIteration:
        workbook.close()
        raise ValueError(f"Workbook had no rows: {path}")
    headers = [safe_text(value) for value in header_values]
    rows: list[dict[str, Any]] = []
    for row_idx, row_values in enumerate(row_iter):
        record: dict[str, Any] = {"row_idx": row_idx, "excel_row": row_idx + 2}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            value = row_values[col_idx] if col_idx < len(row_values) else None
            if value is not None and safe_text(value):
                record[header] = value
        rows.append(record)
    workbook.close()
    return headers, rows


def get_value(row: dict[str, Any], *names: str, limit: int | None = None) -> str:
    for name in names:
        if name in row:
            text = safe_text(row.get(name), limit=limit)
            if text:
                return text
    return ""


def status_is_exact(row: dict[str, Any]) -> bool:
    return get_value(row, "Match Status", "match_status", "Verification Status", "verification_status").lower() == "exact"


def infer_marketplace_from_url(url: str) -> str:
    lowered = url.lower()
    if "amazon." in lowered:
        return "Amazon"
    if "walmart." in lowered:
        return "Walmart"
    if "target." in lowered:
        return "Target"
    if "cvs." in lowered:
        return "CVS"
    if "walgreens." in lowered:
        return "Walgreens"
    if "kroger." in lowered:
        return "Kroger"
    if "ulta." in lowered:
        return "Ulta"
    if "riteaid." in lowered:
        return "Rite Aid"
    return ""


def source_marketplace(row: dict[str, Any]) -> str:
    explicit = get_value(row, "Source_Platform", "Source_Marketplace", "Source_Retailer")
    if explicit:
        return explicit
    inferred = infer_marketplace_from_url(get_value(row, "Source_URL"))
    if inferred:
        return inferred
    if get_value(row, "Source_ASIN"):
        return "Amazon"
    return "Source"


def target_marketplace(row: dict[str, Any]) -> str:
    explicit = get_value(row, "Target_Platform", "Target_Marketplace", "Target_Retailer")
    if explicit:
        return explicit
    inferred = infer_marketplace_from_url(get_value(row, "Target_URL"))
    if inferred:
        return inferred
    return "Target"


def ensure_columns(worksheet: Any, column_names: list[str]) -> dict[str, int]:
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    column_map = {safe_text(header): index for index, header in enumerate(headers, start=1) if safe_text(header)}
    next_col = worksheet.max_column + 1
    for column_name in column_names:
        if column_name not in column_map:
            worksheet.cell(1, next_col).value = column_name
            column_map[column_name] = next_col
            next_col += 1
    return column_map


def extract_json_array(text: str) -> Any:
    stripped = text.strip()
    fenced = re.search(r"```(?:json)?\s*(.*?)\s*```", stripped, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        stripped = fenced.group(1).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        pass
    start = stripped.find("[")
    end = stripped.rfind("]")
    if start < 0 or end <= start:
        raise ValueError("No JSON array found in raw response")
    return json.loads(stripped[start : end + 1])


def observation_key(observation: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        safe_text(observation.get("m")).lower(),
        safe_text(observation.get("id")).lower(),
        safe_text(observation.get("p")),
        safe_text(observation.get("cur")).lower(),
    )
