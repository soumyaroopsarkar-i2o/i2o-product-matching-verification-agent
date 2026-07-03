from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from _paths import DATA_DIR
import shutil

import openpyxl


BASE = DATA_DIR
WORKBOOK = DATA_DIR / "lorealpi_product_verification_output_final.xlsx"
BACKUP = DATA_DIR / "lorealpi_product_verification_output_final.backup-before-walmart-seller-filter.xlsx"
WALMART_JSONL = BASE / "loreal_upc_results" / "loreal_upc_walmart.jsonl"

MAIN_SHEET = "Product Verification"
SUMMARY_SHEET = "Summary"
WALMART_SELLERS = {"walmart", "walmart.com"}


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def normalize_seller(value) -> str:
    return re.sub(r"\s+", " ", clean(value)).lower()


def is_walmart_seller(value) -> bool:
    seller = normalize_seller(value)
    return seller in WALMART_SELLERS or seller.startswith("walmart.com ")


def normalize_upc(value) -> str:
    digits = re.sub(r"\D", "", clean(value))
    return digits.lstrip("0") or digits


def iter_cache_items():
    with WALMART_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            payload = json.loads(line)
            if isinstance(payload, list):
                yield from payload
            else:
                yield payload


def cache_walmart_upcs() -> set[str]:
    upcs: set[str] = set()
    for item in iter_cache_items():
        req = item.get("request") or {}
        result = item.get("result") or {}
        product = result.get("product") if isinstance(result, dict) else {}
        product = product or {}
        buybox = product.get("buybox_winner") if isinstance(product.get("buybox_winner"), dict) else {}
        seller = buybox.get("seller") if isinstance(buybox.get("seller"), dict) else {}
        if is_walmart_seller(seller.get("name")):
            upc = normalize_upc(req.get("gtin") or product.get("upc"))
            if upc:
                upcs.add(upc)
    return upcs


def header_map(ws) -> dict[str, int]:
    return {clean(cell.value): cell.column for cell in ws[1] if clean(cell.value)}


def update_summary(ws, stats: dict[str, int]) -> None:
    if ws is None:
        return

    for row in ws.iter_rows():
        key = clean(row[0].value) if row else ""
        if len(row) < 2:
            continue
        if key == "Total Rows" or key == "Total Rows Reviewed":
            row[1].value = stats["kept_rows"]
        elif "Exact Match" in key or key == "Exact Matches":
            row[1].value = stats["exact_matches"]
        elif "Equivalent" in key:
            row[1].value = stats["equivalent_matches"]
        elif "No Match" in key or "Not Match" in key or key == "Not Matches":
            row[1].value = stats["not_matches"]
        elif key == "Blank / Pending":
            row[1].value = stats["blank_pending"]


def main() -> None:
    cached_walmart_upcs = cache_walmart_upcs()

    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)

    wb = openpyxl.load_workbook(BACKUP)
    ws = wb[MAIN_SHEET]
    headers = header_map(ws)
    seller_col = headers["Target_Seller"]
    upc_col = headers["Source_UPC"]
    status_col = headers["Match Status"]

    original_max_row = ws.max_row
    original_max_col = ws.max_column
    data_rows = []
    upc_rows: dict[str, list[int]] = defaultdict(list)

    for row_idx in range(2, ws.max_row + 1):
        row_cells = [
            (ws.cell(row_idx, col_idx).value, ws.cell(row_idx, col_idx).hyperlink)
            for col_idx in range(1, original_max_col + 1)
        ]
        data_rows.append((row_idx, row_cells))
        upc_rows[normalize_upc(ws.cell(row_idx, upc_col).value)].append(row_idx)

    rows_to_drop: set[int] = set()
    seller_counts = Counter()
    deletion_reasons = Counter()

    for upc, row_indices in upc_rows.items():
        walmart_rows = [
            row_idx
            for row_idx in row_indices
            if is_walmart_seller(ws.cell(row_idx, seller_col).value)
        ]

        for row_idx in row_indices:
            seller = clean(ws.cell(row_idx, seller_col).value)
            seller_counts[seller] += 1

            if row_idx in walmart_rows:
                continue

            if walmart_rows:
                deletion_reasons["removed_non_walmart_when_upc_has_walmart_row"] += 1
            elif upc in cached_walmart_upcs:
                deletion_reasons["removed_non_walmart_cache_has_walmart_seller"] += 1
            else:
                deletion_reasons["removed_non_walmart_no_walmart_seller_in_output"] += 1

            rows_to_drop.add(row_idx)

    kept_rows = [row_cells for row_idx, row_cells in data_rows if row_idx not in rows_to_drop]
    rewrite_rows = kept_rows

    for out_row_idx, row_cells in enumerate(rewrite_rows, start=2):
        for out_col_idx, (value, hyperlink) in enumerate(row_cells, start=1):
            dst = ws.cell(out_row_idx, out_col_idx)
            dst.value = value
            dst.hyperlink = hyperlink.target if hyperlink and hyperlink.target else None

    new_max_row = len(rewrite_rows) + 1
    if new_max_row < original_max_row:
        ws.delete_rows(new_max_row + 1, original_max_row - new_max_row)

    if ws.auto_filter:
        last_col = ws.cell(1, original_max_col).coordinate.replace("1", "")
        ws.auto_filter.ref = f"A1:{last_col}{new_max_row}"

    statuses = Counter(clean(ws.cell(row_idx, status_col).value) for row_idx in range(2, ws.max_row + 1))
    stats = {
        "kept_rows": ws.max_row - 1,
        "exact_matches": statuses["Exact Match"] + statuses["Exact"],
        "equivalent_matches": statuses["Equivalent Match"] + statuses["Equivalent"],
        "not_matches": statuses["Not a Match"] + statuses["Not Match"],
        "blank_pending": statuses[""],
    }
    update_summary(wb[SUMMARY_SHEET] if SUMMARY_SHEET in wb.sheetnames else None, stats)

    wb.save(WORKBOOK)

    print(f"backup={BACKUP}")
    print(f"cached_walmart_upcs={len(cached_walmart_upcs)}")
    print(f"original_rows={sum(seller_counts.values())}")
    print(f"deleted_rows={len(rows_to_drop)}")
    print(f"kept_rows={stats['kept_rows']}")
    print(f"seller_counts={dict(seller_counts.most_common(20))}")
    print(f"deletion_reasons={dict(deletion_reasons)}")
    print(f"statuses={dict(statuses)}")


if __name__ == "__main__":
    main()

