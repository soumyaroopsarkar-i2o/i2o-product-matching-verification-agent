from __future__ import annotations

import csv
from pathlib import Path

from _paths import DATA_DIR

import openpyxl


ROOT = DATA_DIR
I2O_CSV = ROOT / "lorealpi_product_verification_output_final_i2o.csv"
CONSOLIDATED_XLSX = ROOT / "Loreal_Consolidated_with_ASIN.xlsx"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    with I2O_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        i2o_rows = list(csv.DictReader(f))
    i2o_asins = [clean(row.get("product_code")) for row in i2o_rows if clean(row.get("product_code"))]

    wb = openpyxl.load_workbook(CONSOLIDATED_XLSX, read_only=True, data_only=True)
    asin_values: set[str] = set()
    any_cell_values: set[str] = set()
    asin_columns: list[tuple[str, list[str]]] = []
    sheet_headers: list[tuple[str, int, int, list[str]]] = []

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            continue

        sheet_headers.append((ws.title, ws.max_row, ws.max_column, headers))
        asin_col_indexes = [idx for idx, header in enumerate(headers) if "asin" in header.lower()]
        if asin_col_indexes:
            asin_columns.append((ws.title, [headers[idx] for idx in asin_col_indexes]))

        for row in rows:
            for value in row:
                text = clean(value)
                if text:
                    any_cell_values.add(text)
            for idx in asin_col_indexes:
                value = clean(row[idx])
                if value:
                    asin_values.add(value)

    wb.close()

    missing = sorted(set(i2o_asins) - asin_values)
    missing_any_cell = sorted(set(i2o_asins) - any_cell_values)

    print(f"i2o_rows={len(i2o_rows)}")
    print(f"i2o_unique_asins={len(set(i2o_asins))}")
    print(f"sheet_headers={sheet_headers}")
    print(f"consolidated_asin_columns={asin_columns}")
    print(f"consolidated_unique_asins={len(asin_values)}")
    print(f"missing_unique_asins={len(missing)}")
    if missing:
        print("missing_examples=" + ", ".join(missing[:50]))
    print(f"found_exact_product_code_asins_any_cell={len(set(i2o_asins) - set(missing_any_cell))}")
    print(f"missing_exact_product_code_asins_any_cell={len(missing_any_cell)}")
    if missing_any_cell:
        print("missing_any_cell_examples=" + ", ".join(missing_any_cell[:50]))


if __name__ == "__main__":
    main()

