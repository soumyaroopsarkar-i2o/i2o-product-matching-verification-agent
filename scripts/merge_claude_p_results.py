from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


OUTPUT_COLUMNS = ("Match Status", "Match Justification")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def ensure_output_columns(worksheet: Any) -> tuple[int, int]:
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    column_map = {str(header).strip(): index for index, header in enumerate(headers, start=1) if header}

    next_col = worksheet.max_column + 1
    output_indices: list[int] = []
    for column_name in OUTPUT_COLUMNS:
        existing = column_map.get(column_name)
        if existing:
            output_indices.append(existing)
        else:
            worksheet.cell(1, next_col).value = column_name
            output_indices.append(next_col)
            next_col += 1
    return output_indices[0], output_indices[1]


def collect_decisions(manifest: dict[str, Any], *, allow_partial: bool) -> tuple[dict[int, dict[str, str]], list[int]]:
    decisions_by_row: dict[int, dict[str, str]] = {}
    missing_batches = []

    for batch in manifest["batches"]:
        result_path = Path(batch["result_path"])
        batch_index = batch["batch_index"]
        if not result_path.exists():
            missing_batches.append(batch_index)
            continue
        result = load_json(result_path)
        for decision in result["decisions"]:
            row_idx = int(decision["row_idx"])
            if row_idx in decisions_by_row:
                raise ValueError(f"Duplicate row_idx across result files: {row_idx}")
            decisions_by_row[row_idx] = {
                "status": str(decision["status"]),
                "justification": str(decision.get("justification") or ""),
            }

    if missing_batches and not allow_partial:
        formatted = ", ".join(f"{index:03d}" for index in missing_batches[:20])
        suffix = "..." if len(missing_batches) > 20 else ""
        raise ValueError(f"Missing result files for batches: {formatted}{suffix}")

    return decisions_by_row, missing_batches


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge validated Claude -p batch results into the final workbook.")
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output-workbook", type=Path, help="Override final output workbook path.")
    parser.add_argument("--allow-partial", action="store_true", help="Write available rows even if some batches are missing.")
    args = parser.parse_args()

    run_dir = args.run_dir.resolve()
    manifest_path = run_dir / "manifest.json"
    manifest = load_json(manifest_path)

    input_workbook = Path(manifest["input_workbook"])
    output_workbook = args.output_workbook or Path(manifest["output_workbook"])
    output_workbook.parent.mkdir(parents=True, exist_ok=True)

    decisions_by_row, missing_batches = collect_decisions(manifest, allow_partial=args.allow_partial)

    expected_rows = {
        row_idx
        for batch in manifest["batches"]
        for row_idx in batch["expected_row_idx"]
    }
    missing_rows = sorted(expected_rows - set(decisions_by_row))
    if missing_rows and not args.allow_partial:
        raise ValueError(f"Missing decisions for row_idx values: {missing_rows[:30]}")

    workbook = openpyxl.load_workbook(input_workbook)
    worksheet = workbook.active
    status_col, justification_col = ensure_output_columns(worksheet)

    for row_idx, decision in decisions_by_row.items():
        excel_row = row_idx + 2
        worksheet.cell(excel_row, status_col).value = decision["status"]
        worksheet.cell(excel_row, justification_col).value = decision["justification"]

    workbook.save(output_workbook)
    workbook.close()

    status_counts = Counter(decision["status"] for decision in decisions_by_row.values())
    summary = {
        "run_id": manifest["run_id"],
        "merged_at": datetime.now().isoformat(timespec="seconds"),
        "input_workbook": str(input_workbook),
        "output_workbook": str(output_workbook),
        "allow_partial": args.allow_partial,
        "total_workbook_data_rows": manifest["total_workbook_data_rows"],
        "selected_data_rows": manifest["selected_data_rows"],
        "completed_rows": len(decisions_by_row),
        "missing_rows": missing_rows,
        "missing_batches": missing_batches,
        "status_counts": dict(status_counts),
    }
    summary_path = output_workbook.with_suffix(".summary.json")
    write_json(summary_path, summary)

    print(f"Wrote workbook: {output_workbook}")
    print(f"Wrote summary: {summary_path}")
    print(f"Completed rows: {len(decisions_by_row)}")
    print(f"Status counts: {dict(status_counts)}")


if __name__ == "__main__":
    main()
