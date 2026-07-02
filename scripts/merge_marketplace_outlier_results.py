from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from price_workflow_common import ensure_columns, load_json, safe_text, write_json


OUTPUT_COLUMNS = [
    "marketplace_having_anomaly",
    "price anomaly justification",
]


def collect_decisions(manifest: dict[str, Any], *, allow_partial: bool) -> tuple[dict[int, dict[str, str]], list[int]]:
    decisions_by_row: dict[int, dict[str, str]] = {}
    missing_batches = []
    for skipped in manifest.get("skipped_decisions", []):
        row_idx = int(skipped["row_idx"])
        decisions_by_row[row_idx] = {
            "marketplace": "",
            "reason": safe_text(skipped.get("reason") or ""),
        }

    for batch in manifest["batches"]:
        result_path = Path(batch["result_path"])
        batch_index = batch["batch_index"]
        if not result_path.exists():
            missing_batches.append(batch_index)
            continue
        result = load_json(result_path)
        for decision in result["decisions"]:
            row_idx = int(decision["row_idx"])
            if row_idx in decisions_by_row:
                raise ValueError(f"Duplicate row_idx across Stage 2B result files: {row_idx}")
            decisions_by_row[row_idx] = {
                "marketplace": safe_text(decision.get("marketplace") or ""),
                "reason": safe_text(decision.get("reason") or ""),
            }

    if missing_batches and not allow_partial:
        formatted = ", ".join(f"{index:03d}" for index in missing_batches[:20])
        suffix = "..." if len(missing_batches) > 20 else ""
        raise ValueError(f"Missing Stage 2B result files for batches: {formatted}{suffix}")

    return decisions_by_row, missing_batches


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge Stage 2B marketplace outlier results into a workbook.")
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output-workbook", type=Path, help="Override Stage 2B output workbook path.")
    parser.add_argument("--allow-partial", action="store_true")
    args = parser.parse_args()

    run_dir = args.run_dir.resolve()
    manifest = load_json(run_dir / "manifest.json")
    input_workbook = Path(manifest["input_workbook"])
    output_workbook = args.output_workbook or Path(manifest["output_workbook"])
    output_workbook.parent.mkdir(parents=True, exist_ok=True)

    decisions_by_row, missing_batches = collect_decisions(manifest, allow_partial=args.allow_partial)
    expected_rows = {
        row_idx
        for batch in manifest["batches"]
        for row_idx in batch["expected_row_idx"]
    }
    missing_rows = sorted(expected_rows - set(decisions_by_row))
    if missing_rows and not args.allow_partial:
        raise ValueError(f"Missing Stage 2B decisions for row_idx values: {missing_rows[:30]}")

    workbook = openpyxl.load_workbook(input_workbook)
    worksheet = workbook.active
    column_map = ensure_columns(worksheet, OUTPUT_COLUMNS)

    status_counts: Counter[str] = Counter()
    for excel_row in range(2, worksheet.max_row + 1):
        worksheet.cell(excel_row, column_map["marketplace_having_anomaly"]).value = ""
        worksheet.cell(excel_row, column_map["price anomaly justification"]).value = ""

    for row_idx, decision in decisions_by_row.items():
        excel_row = row_idx + 2
        marketplace = decision["marketplace"]
        worksheet.cell(excel_row, column_map["marketplace_having_anomaly"]).value = marketplace
        worksheet.cell(excel_row, column_map["price anomaly justification"]).value = decision["reason"]
        status_counts["Marketplace identified" if marketplace else "No marketplace identified"] += 1

    workbook.save(output_workbook)
    workbook.close()

    summary = {
        "stage": manifest["stage"],
        "run_id": manifest["run_id"],
        "merged_at": datetime.now().isoformat(timespec="seconds"),
        "input_workbook": str(input_workbook),
        "output_workbook": str(output_workbook),
        "allow_partial": args.allow_partial,
        "candidate_reclassified_rows": manifest["candidate_reclassified_rows"],
        "candidate_stage1_equivalent_rows": manifest.get("candidate_stage1_equivalent_rows", 0),
        "candidate_equivalent_rows": manifest.get("candidate_equivalent_rows", manifest["candidate_reclassified_rows"]),
        "eligible_rows": manifest["eligible_rows"],
        "completed_rows": len(decisions_by_row),
        "missing_rows": missing_rows,
        "missing_batches": missing_batches,
        "status_counts": dict(status_counts),
    }
    summary_path = output_workbook.with_suffix(".summary.json")
    write_json(summary_path, summary)

    print(f"Wrote Stage 2B workbook: {output_workbook}")
    print(f"Wrote Stage 2B summary: {summary_path}")
    print(f"Completed Stage 2B rows: {len(decisions_by_row)}")
    print(f"Status counts: {dict(status_counts)}")


if __name__ == "__main__":
    main()
