from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from price_workflow_common import ensure_columns, load_json, safe_text, write_json


OUTPUT_COLUMNS = [
    "reclassified_status",
]

EQUIV_CODES = {"E", "EQ", "EQUIV", "EQUIVALENT"}


def collect_decisions(manifest: dict[str, Any], *, allow_partial: bool) -> tuple[dict[int, dict[str, str]], list[int]]:
    decisions_by_row: dict[int, dict[str, str]] = {}
    missing_batches = []
    for batch in manifest["batches"]:
        result_path = Path(batch["result_path"])
        batch_index = batch["batch_index"]
        if not result_path.exists():
            missing_batches.append(batch_index)
            continue
        result = load_json(result_path)
        for decision in result["decisions"]:
            row_idx = int(decision["row_idx"])
            status_code = safe_text(decision.get("status_code", "EQUIV")).upper()
            if status_code not in EQUIV_CODES:
                continue
            if row_idx in decisions_by_row:
                raise ValueError(f"Duplicate row_idx across Stage 2A result files: {row_idx}")
            decisions_by_row[row_idx] = {
                "reclassified_status": "Equivalent",
            }

    if missing_batches and not allow_partial:
        formatted = ", ".join(f"{index:03d}" for index in missing_batches[:20])
        suffix = "..." if len(missing_batches) > 20 else ""
        raise ValueError(f"Missing Stage 2A result files for batches: {formatted}{suffix}")

    return decisions_by_row, missing_batches


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge Stage 2A pairwise price anomaly results into a workbook.")
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output-workbook", type=Path, help="Override Stage 2A output workbook path.")
    parser.add_argument("--allow-partial", action="store_true")
    args = parser.parse_args()

    run_dir = args.run_dir.resolve()
    manifest = load_json(run_dir / "manifest.json")
    input_workbook = Path(manifest["input_workbook"])
    output_workbook = args.output_workbook or Path(manifest["output_workbook"])
    output_workbook.parent.mkdir(parents=True, exist_ok=True)

    decisions_by_row, missing_batches = collect_decisions(manifest, allow_partial=args.allow_partial)
    workbook = openpyxl.load_workbook(input_workbook)
    worksheet = workbook.active
    column_map = ensure_columns(worksheet, OUTPUT_COLUMNS)
    status_counts: Counter[str] = Counter()
    reclassified_count = 0

    for excel_row in range(2, worksheet.max_row + 1):
        row_idx = excel_row - 2
        decision = decisions_by_row.get(row_idx)
        reclassified = decision["reclassified_status"] if decision else ""
        worksheet.cell(excel_row, column_map["reclassified_status"]).value = reclassified
        if reclassified == "Equivalent":
            reclassified_count += 1
            status_counts["Equivalent"] += 1

    workbook.save(output_workbook)
    workbook.close()

    summary = {
        "stage": manifest["stage"],
        "run_id": manifest["run_id"],
        "merged_at": datetime.now().isoformat(timespec="seconds"),
        "input_workbook": str(input_workbook),
        "output_workbook": str(output_workbook),
        "allow_partial": args.allow_partial,
        "selected_exact_rows": manifest["selected_exact_rows"],
        "reclassified_rows": len(decisions_by_row),
        "reclassified_equivalent_rows": reclassified_count,
        "missing_batches": missing_batches,
        "status_counts": dict(status_counts),
    }
    summary_path = output_workbook.with_suffix(".summary.json")
    write_json(summary_path, summary)

    print(f"Wrote Stage 2A workbook: {output_workbook}")
    print(f"Wrote Stage 2A summary: {summary_path}")
    print(f"Stage 2A reclassified rows: {len(decisions_by_row)}")
    print(f"Reclassified Exact -> Equivalent rows: {reclassified_count}")


if __name__ == "__main__":
    main()
