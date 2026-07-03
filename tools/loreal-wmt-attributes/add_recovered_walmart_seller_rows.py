from __future__ import annotations

import json
import re
from collections import Counter
from copy import copy
from pathlib import Path

from _paths import DATA_DIR

import openpyxl


BASE = DATA_DIR
ROOT = DATA_DIR
WORKBOOK = ROOT / "lorealpi_product_verification_output_final.xlsx"
ALT_WORKBOOK = ROOT / "lorealpi_product_verification_output_final_with_recovered_walmart.xlsx"
BACKUP = ROOT / "lorealpi_product_verification_output_final.backup-before-walmart-seller-filter.xlsx"
RECOVERY_JSONL = BASE / "removed_upcs_bluecart_search_cache.jsonl"

MAIN_SHEET = "Product Verification"
SUMMARY_SHEET = "Summary"


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def norm_upc(value) -> str:
    digits = re.sub(r"\D", "", clean(value))
    return digits.lstrip("0") or digits


def is_walmart(value) -> bool:
    seller = clean(value).lower()
    return seller in {"walmart", "walmart.com"} or seller.startswith("walmart.com ")


def seller_name(obj) -> str:
    if isinstance(obj, dict):
        return clean(obj.get("name") or obj.get("seller") or obj.get("displayName"))
    return clean(obj)


def category_path(product: dict) -> str:
    crumbs = product.get("breadcrumbs") or product.get("categories") or []
    if isinstance(crumbs, list):
        return " > ".join(clean(c.get("name")) for c in crumbs if isinstance(c, dict) and clean(c.get("name")))
    return ""


def main_image(product: dict) -> str:
    main = product.get("main_image")
    if isinstance(main, dict):
        return clean(main.get("link"))
    if isinstance(main, str):
        return clean(main)
    images = product.get("images") or []
    if images and isinstance(images[0], dict):
        return clean(images[0].get("link"))
    if images and isinstance(images[0], str):
        return clean(images[0])
    return ""


def recovered_products() -> dict[str, dict]:
    out = {}
    if not RECOVERY_JSONL.exists():
        return out
    with RECOVERY_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            record = json.loads(line)
            upc = norm_upc(record.get("upc"))
            for payload in record.get("product_payloads") or []:
                product = payload.get("product") or {}
                buybox = product.get("buybox_winner") if isinstance(product.get("buybox_winner"), dict) else {}
                seller = seller_name(buybox.get("seller"))
                if is_walmart(seller) and norm_upc(product.get("upc")) == upc:
                    out.setdefault(upc, product)
                    break
    return out


def header_map(ws) -> dict[str, int]:
    return {clean(cell.value): cell.column for cell in ws[1] if clean(cell.value)}


def load_backup_rows() -> dict[str, dict]:
    wb = openpyxl.load_workbook(BACKUP, read_only=True, data_only=True)
    ws = wb[MAIN_SHEET]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    out = {}
    for values in rows:
        row = dict(zip(headers, values))
        upc = norm_upc(row.get("Source_UPC"))
        seller = clean(row.get("Target_Seller"))
        if upc and not is_walmart(seller):
            out.setdefault(upc, row)
    return out


def set_cell(ws, row_idx: int, headers: dict[str, int], name: str, value) -> None:
    if name in headers:
        ws.cell(row_idx, headers[name]).value = value


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col_idx)
        dst = ws.cell(dst_row, col_idx)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def update_summary(ws, stats: Counter) -> None:
    for row in ws.iter_rows():
        key = clean(row[0].value) if row else ""
        if len(row) < 2:
            continue
        if key == "Total Rows" or key == "Total Rows Reviewed":
            row[1].value = stats["total"]
        elif "Exact Match" in key or key == "Exact Matches":
            row[1].value = stats["exact"]
        elif "Equivalent" in key:
            row[1].value = stats["equivalent"]
        elif "No Match" in key or "Not Match" in key or key == "Not Matches":
            row[1].value = stats["not_match"]


def workbook_stats(ws, headers: dict[str, int]) -> Counter:
    status_col = headers["Match Status"]
    stats = Counter(total=ws.max_row - 1)
    for row_idx in range(2, ws.max_row + 1):
        status = clean(ws.cell(row_idx, status_col).value).lower()
        if status in {"exact", "exact match"}:
            stats["exact"] += 1
        elif "equivalent" in status:
            stats["equivalent"] += 1
        elif status in {"not a match", "not match", "no match"}:
            stats["not_match"] += 1
    return stats


def main() -> None:
    recovered = recovered_products()
    backup_rows = load_backup_rows()

    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[MAIN_SHEET]
    headers = header_map(ws)
    current_upcs = {
        norm_upc(ws.cell(row_idx, headers["Source_UPC"]).value)
        for row_idx in range(2, ws.max_row + 1)
    }

    appended = 0
    for upc, product in recovered.items():
        if upc in current_upcs or upc not in backup_rows:
            continue

        src = backup_rows[upc]
        buybox = product.get("buybox_winner") if isinstance(product.get("buybox_winner"), dict) else {}
        seller = seller_name(buybox.get("seller")) or "Walmart.com"
        row_idx = ws.max_row + 1
        copy_row_style(ws, 2, row_idx)

        for name, col_idx in headers.items():
            ws.cell(row_idx, col_idx).value = src.get(name)

        set_cell(ws, row_idx, headers, "Target_Platform", "walmart.com")
        set_cell(ws, row_idx, headers, "Target_Item_ID", clean(product.get("item_id")))
        set_cell(ws, row_idx, headers, "Target_Product_ID", clean(product.get("product_id")))
        set_cell(ws, row_idx, headers, "Target_URL", clean(product.get("link")))
        set_cell(ws, row_idx, headers, "Target_Brand", clean(product.get("brand")))
        set_cell(ws, row_idx, headers, "Target_Title", clean(product.get("title")))
        set_cell(ws, row_idx, headers, "Target_Model", clean(product.get("model")))
        set_cell(ws, row_idx, headers, "Target_Type", clean(product.get("type")))
        set_cell(ws, row_idx, headers, "Target_Category", category_path(product))
        set_cell(ws, row_idx, headers, "Target_Rating", product.get("rating"))
        set_cell(ws, row_idx, headers, "Target_Ratings_Total", product.get("ratings_total"))
        set_cell(ws, row_idx, headers, "Target_Seller", seller)
        set_cell(ws, row_idx, headers, "Target_Description", clean(product.get("description")))
        set_cell(ws, row_idx, headers, "Target_Ingredients", clean(product.get("ingredients")))
        set_cell(ws, row_idx, headers, "Target_Image_URL", main_image(product))
        set_cell(ws, row_idx, headers, "Match Status", "Exact")
        set_cell(
            ws,
            row_idx,
            headers,
            "Match Justification",
            "Recovered via live BlueCart lookup: exact UPC match and Walmart.com buybox seller.",
        )

        for link_col in ("Source_URL", "Source_Image_URL", "Target_URL", "Target_Image_URL"):
            if link_col in headers:
                cell = ws.cell(row_idx, headers[link_col])
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

        current_upcs.add(upc)
        appended += 1

    if appended:
        if ws.auto_filter:
            last_col = ws.cell(1, ws.max_column).coordinate.replace("1", "")
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
        if SUMMARY_SHEET in wb.sheetnames:
            update_summary(wb[SUMMARY_SHEET], workbook_stats(ws, headers))
        try:
            wb.save(WORKBOOK)
            saved_to = WORKBOOK
        except PermissionError:
            wb.save(ALT_WORKBOOK)
            saved_to = ALT_WORKBOOK
    else:
        saved_to = WORKBOOK

    print(f"recovered_exact_walmart_upcs={len(recovered)}")
    print(f"appended_rows={appended}")
    print(f"workbook_rows={ws.max_row - 1}")
    print(f"saved_to={saved_to}")


if __name__ == "__main__":
    main()

